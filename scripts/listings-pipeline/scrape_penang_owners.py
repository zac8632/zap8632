#!/usr/bin/env python3
"""
Penang property scraper (For Sale + For Rent, By Owner only) for mudah.my,
covering residential, commercial, land and industrial listings.

Uses curl_cffi with Chrome TLS-fingerprint impersonation to fetch listing
search pages exactly like a real browser at the network level, then reads
the __NEXT_DATA__ JSON embedded in every mudah.my search results page
directly - this sidesteps the masked-phone-number / click-to-reveal
problems seen when driving a headless browser: the phone number is present
as plain JSON in the page payload, no button click, no login, no partial
masking.

Output columns match what was asked for from the live site (listing card +
description panel):
    Title (product/project name + area), Category (e.g. "Condominium For
    Sale"), Listed Date (DD/MM/YYYY), Price (RM x,xxx,xxx), Location, Area
    size (built-up + land), Bedrooms/Bathrooms, Tenure, Furnishing,
    Description, Seller Name, Agency, Phone, Has WhatsApp, Listing URL.

"By Owner" filtering: mudah.my's own UI has a filter tab for this but this
script doesn't call a confirmed query param for it (see scraper.py's
switch_to_owner_only for the click-based approach against the UI instead).
This script instead replicates the same distinction client-side: an ad
counts as "by owner" if it has no agency/company name attached to it
(agentData empty/missing), which is what live testing showed private
sellers looking like ("Private advertiser" label, no agency) versus agent
listings (agency name present). Use --include-agents to disable this
filter and keep everything.

Install:
    pip install curl_cffi pandas openpyxl

Usage:
    python scrape_penang_owners.py --pages 50 --output penang_owners.xlsx
    python scrape_penang_owners.py --pages 2 --debug-dump   # inspect field names
    python scrape_penang_owners.py --statewide              # all of Penang, not just the island
    python scrape_penang_owners.py --fetch-details          # also visit each listing page for
                                                              # the full Description + any detail
                                                              # fields missing from the search JSON

Detail-page enrichment (--fetch-details): the search-results __NEXT_DATA__
payload only carries summary fields for each card - the full multi-paragraph
Description, and sometimes Tenure/Furnishing/exact
Bedrooms/Bathrooms/Size, live only on the listing's own detail page. With
--fetch-details, this script makes one extra request per listing (same
curl_cffi + __NEXT_DATA__ technique, no browser) and fills in whatever
columns came back empty from the search page. This roughly doubles the
number of requests and run time, so it's opt-in; use --detail-limit to cap
how many listings get this treatment (0 = all).
"""

import argparse
import datetime
import json
import re
import sys
import time
import random

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Penang Island subarea IDs (George Town / Tanjong Tokong / Tanjung Bungah /
# Pulau Tikus / etc.) - default scope throughout this project, taken from
# the user-supplied mudah.my search URL. Pass --statewide to scrape all of
# Penang state instead (e.g. https://www.mudah.my/penang/properties-for-sale,
# ~1858 "By Owner" listings as of the last check) rather than just the island.
PENANG_SUBAREAS = "117,116,102,99,88"

# (label, URL slug, asset_type). "sale"/"rent" residential slugs are
# confirmed working (live-tested). commercial/land slugs follow mudah.my's
# usual naming convention but are NOT yet live-verified - run with
# --debug-dump on these first and check for 0-listing results, which would
# mean the slug is wrong. Mudah doesn't appear to have a distinct top-level
# "industrial" category; industrial listings are expected to show up under
# commercial with their own subCategory field instead (see Property Type
# column) - verify this once real data is seen.
CATEGORIES = [
    ("sale-residential",   "properties-for-sale",             "residential"),
    ("rent-residential",   "properties-for-rent",             "residential"),
    ("sale-commercial",    "commercial-properties-for-sale",  "commercial"),
    ("rent-commercial",    "commercial-properties-for-rent",  "commercial"),
    ("sale-land",          "land-for-sale",                   "land"),
]

# Hero image - the search-results ad JSON already embeds the card thumbnail,
# same Apollo CDN pattern build_listing_posts.py upsizes to full-res for
# qualifying listings. Grabbing just the first one here costs zero extra
# requests and is enough to recognise the unit from the Sheet/Airtable.
_BAD_IMG = ("sprite", "logo", "icon", ".svg", "placeholder", "avatar", "favicon", "flag", "sprites")
_APOLLO_SIZE_RE = re.compile(r';s=\d+x\d+', re.IGNORECASE)


def _is_photo_url(u):
    ul = u.lower()
    if any(b in ul for b in _BAD_IMG):
        return False
    return ("apollo" in ul or "akamaized" in ul or ";s=" in ul
            or re.search(r"\.(?:jpg|jpeg|png|webp)(?:\?|;|$)", ul))


def _all_strings(obj, out, depth=0):
    if depth > 14:
        return
    if isinstance(obj, str):
        if obj.startswith("http"):
            out.append(obj)
    elif isinstance(obj, dict):
        for v in obj.values():
            _all_strings(v, out, depth + 1)
    elif isinstance(obj, list):
        for v in obj:
            _all_strings(v, out, depth + 1)


def extract_hero_image_url(item):
    """First real photo URL straight from the search-results ad data - a
    thumbnail-sized rendition, just enough to identify the unit. Full-res
    photo fetching for qualifying listings still happens in Stage 2."""
    strs = []
    _all_strings(item, strs)
    photos = [u for u in strs if _is_photo_url(u)]
    if not photos:
        return None
    return _APOLLO_SIZE_RE.sub(";s=400x300", photos[0])


# Field names this script tries for listing details, in priority order.
# mudah.my's ad payload shape has varied across scrapes in this project -
# if these come up empty, check --debug-dump output for the real key names.
def classify_asset_type(property_type, default):
    """Mudah's general properties-for-sale/for-rent search already mixes in
    commercial, land, and room listings alongside residential ones - the
    asset_type passed down from CATEGORIES only reflects which URL was
    queried, not what a given listing actually is. Derive the real type
    from its own Property Type text instead, falling back to the queried
    category\'s asset_type when Property Type is missing."""
    if not property_type:
        return default
    pt = str(property_type).lower()
    if "land" in pt:
        return "land"
    if any(k in pt for k in ("commercial", "office", "shop", "retail", "industrial", "factory", "warehouse", "shoplot", "showroom")):
        return "commercial"
    return "residential"


MUDAH_FIELD_CANDIDATES = {
    "project_name": ["projectName", "developmentName", "buildingName"],
    "headline":     ["subject", "title", "adTitle"],
    "description":  ["body", "description", "adDescription", "content", "descriptionText", "adBody"],
    "price":        ["price", "priceLabel"],
    "location":     ["subareaName", "locationLabel", "areaName"],
    "state":        ["regionName", "stateName"],
    "property_type":["subCategoryName", "propertyType", "categoryName", "subCategory"],
    "tenure":       ["tenure", "attr_tenure"],
    "furnishing":   ["furnishing", "attr_furnishing"],
    "bedroom":      ["bedroom", "attr_bedroom", "bedrooms"],
    "bathroom":     ["bathroom", "attr_bathroom", "bathrooms"],
    "size":         ["size", "attr_size", "builtUpSize"],
    "land_size":    ["landSize", "attr_land_size", "lotSize"],
    "listed_at":    ["publishedDatetime", "dateAdded", "mDate", "createdAt"],
}

# Columns that are worth filling in from a listing's own detail page when
# --fetch-details is used and the search-results JSON left them empty.
# Maps output column name -> MUDAH_FIELD_CANDIDATES key.
DETAIL_FILL_COLUMNS = {
    "Description":   "description",
    "Tenure":        "tenure",
    "Furnishing":    "furnishing",
    "Bedrooms":      "bedroom",
    "Bathrooms":     "bathroom",
    "Size (sqft)":   "size",
    "Land Size":     "land_size",
    "Property Type": "property_type",
    "Listed Date":   "listed_at",
}

# Tenure/Furnishing/Bedrooms/Bathrooms/etc mostly aren't plain top-level
# attributes - a real debug dump showed them living inside a
# "categoryParams" list (each item like {"id": "rooms", "value": "3",
# "label": "Bedrooms"}) and/or a "propertyParams" list of {header,
# params: [{id, value, ...}]} groups. Maps MUDAH_FIELD_CANDIDATES key ->
# the param "id" values to look for in those lists, checked as a fallback
# whenever the direct top-level field lookup above comes back empty.
PARAM_ID_CANDIDATES = {
    "tenure":        ["title_type", "tenure"],
    "furnishing":    ["furnishing"],
    "bedroom":       ["rooms", "bedroom"],
    "bathroom":      ["bathroom"],
    "property_type": ["property_type"],
    "size":          ["size"],
    "land_size":     ["land_size", "lot_size"],
}


def extract_category_params(attrs):
    """Flatten an ad's categoryParams / propertyParams lists into a plain
    {param_id: value} dict for easy fallback lookups."""
    params = {}
    if not isinstance(attrs, dict):
        return params
    for item in attrs.get("categoryParams") or []:
        if isinstance(item, dict) and item.get("id"):
            params.setdefault(item["id"], item.get("value"))
    for group in attrs.get("propertyParams") or []:
        if isinstance(group, dict):
            for item in group.get("params") or []:
                if isinstance(item, dict) and item.get("id"):
                    params.setdefault(item["id"], item.get("value"))
    return params


def field_value(attrs, key, params=None):
    """Look up a MUDAH_FIELD_CANDIDATES field, falling back to the
    categoryParams/propertyParams dict (see PARAM_ID_CANDIDATES) when the
    direct top-level field isn't present."""
    val = _first(attrs, MUDAH_FIELD_CANDIDATES[key])
    if val:
        return val
    if params is None:
        params = extract_category_params(attrs)
    return _first(params, PARAM_ID_CANDIDATES.get(key, []))


def _first(d, keys):
    for k in keys:
        v = d.get(k)
        if v not in (None, ""):
            return v
    return None


def clean_description(text):
    """mudah.my's description field ('body') comes as raw HTML - just
    <br> line breaks in practice, but strip any other stray tags too so
    Excel shows readable text instead of markup."""
    if not text:
        return text
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


PHONE_DIGITS_RE = re.compile(r"^1[0-46-9]\d{7,8}$")


def norm_phone(raw):
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if digits.startswith("60"):
        digits = digits[2:]
    elif digits.startswith("0"):
        digits = digits[1:]
    if not PHONE_DIGITS_RE.match(digits):
        return None
    return "0" + digits


def polite_sleep(a=2.0, b=4.0):
    time.sleep(random.uniform(a, b))


def is_owner_listing(inner):
    """An ad counts as 'by owner' if it has no agency/company name attached.
    See module docstring for why this heuristic was chosen over a query
    param."""
    if not isinstance(inner, dict):
        return True
    agency = inner.get("storeParamsCompanyName") or inner.get("storeName")
    return not bool(agency)


def format_date_ddmmyyyy(value):
    """Convert whatever date shape mudah.my's JSON gives us (epoch
    seconds/ms, ISO string, or relative text like "3 days ago") into
    DD/MM/YYYY. Falls back to the raw value (or None) if nothing matches."""
    if value in (None, ""):
        return None

    # Numeric epoch timestamp (seconds or milliseconds).
    if isinstance(value, (int, float)) or (isinstance(value, str) and re.match(r"^\d{9,13}$", value.strip())):
        try:
            n = float(value)
            if n > 1e12:  # milliseconds
                n /= 1000
            return datetime.datetime.utcfromtimestamp(n).strftime("%d/%m/%Y")
        except Exception:
            pass

    text = str(value).strip()

    # ISO-ish date string, e.g. "2026-06-30T12:00:00Z" or "2026-06-30".
    iso_m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if iso_m:
        y, mo, d = iso_m.groups()
        try:
            return datetime.date(int(y), int(mo), int(d)).strftime("%d/%m/%Y")
        except Exception:
            pass

    # Already DD/MM/YYYY or similar - leave as-is.
    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", text):
        parts = text.split("/")
        return f"{int(parts[0]):02d}/{int(parts[1]):02d}/{parts[2]}"

    # Relative text: "today", "yesterday", "X days/weeks/months ago" - also
    # matches mudah.my's own "Today HH:MM" / "Yesterday HH:MM" format
    # (confirmed live: publishedDatetime = "Today 13:58").
    today = datetime.date.today()
    low = text.lower()
    if low == "today" or low.startswith("today "):
        return today.strftime("%d/%m/%Y")
    if low == "yesterday" or low.startswith("yesterday "):
        return (today - datetime.timedelta(days=1)).strftime("%d/%m/%Y")
    rel_m = re.match(r"^(\d+)\s*(day|days|week|weeks|month|months|hour|hours|minute|minutes)\s*ago$", low)
    if rel_m:
        n = int(rel_m.group(1))
        unit = rel_m.group(2)
        if unit.startswith("hour") or unit.startswith("minute"):
            d = today
        elif unit.startswith("day"):
            d = today - datetime.timedelta(days=n)
        elif unit.startswith("week"):
            d = today - datetime.timedelta(weeks=n)
        else:  # month(s) - approximate as 30 days
            d = today - datetime.timedelta(days=30 * n)
        return d.strftime("%d/%m/%Y")

    # "DD Mon YYYY" style, e.g. "17 Jun 2025" (older listings), with or
    # without a trailing time.
    for fmt in ("%d %b %Y %H:%M", "%d %b %Y"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass

    # Unrecognised format - return the raw text rather than losing the data.
    return text


def format_price(raw):
    """Format a raw price value (numeric or string) as "RM x,xxx,xxx"."""
    if raw in (None, ""):
        return None
    text = str(raw).strip()
    digits = re.sub(r"[^\d.]", "", text)
    if not digits:
        return text if text else None
    try:
        n = float(digits)
        if n == int(n):
            return f"RM {int(n):,}"
        return f"RM {n:,.2f}"
    except Exception:
        return text


def build_title(a):
    """Title = "{Project/Product Name}, {Area}" - the condo/product name
    line from the listing card plus its area, not the ad's custom headline
    (that goes in Description instead)."""
    project = _first(a, MUDAH_FIELD_CANDIDATES["project_name"])
    area = _first(a, MUDAH_FIELD_CANDIDATES["location"])
    if project and area:
        return f"{project}, {area}"
    if project:
        return project
    # No distinct project-name field found in the payload - fall back to
    # the headline (still better than a blank Title) so nothing is lost.
    headline = _first(a, MUDAH_FIELD_CANDIDATES["headline"])
    if headline and area:
        return f"{headline}, {area}"
    return headline or area


def build_category(label, a):
    """e.g. "Condominium For Sale" - property type + For Sale/For Rent."""
    property_type = field_value(a, "property_type")
    action = "For Rent" if label.startswith("rent") else "For Sale"
    if property_type:
        return f"{property_type} {action}"
    return action


def slugify(text):
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def build_listing_url(a):
    """A bare 'https://www.mudah.my/{listId}' 404s. Confirmed live pattern
    (from a real listing link the user pasted back):
    'https://www.mudah.my/{slugified-headline}-{listId}.htm' - e.g. headline
    "Looking For Investors Only" + listId 115155444 ->
    'https://www.mudah.my/looking-for-investors-only-115155444.htm'. Built
    directly from the headline rather than searched for in the JSON, since
    the slug is deterministic from the title text."""
    list_id = str(a.get("listId", ""))
    if not list_id:
        return ""
    headline = _first(a, MUDAH_FIELD_CANDIDATES["headline"])
    if headline:
        return f"https://www.mudah.my/{slugify(headline)}-{list_id}.htm"
    # No headline available - bare ID URL as last resort (known to 404 on
    # direct visit, but still useful as a stable per-listing identifier).
    return f"https://www.mudah.my/{list_id}"


NEXT_DATA_RE = re.compile(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', re.DOTALL)


def find_ad_node(obj, list_id, _depth=0):
    """Recursively search a listing detail page's __NEXT_DATA__ tree for the
    dict describing this specific ad, matched by listId. Detail pages don't
    share one confirmed JSON shape with the search-results payload (and may
    differ between residential/commercial/land templates), so rather than
    guess a fixed path this walks the whole tree looking for a node whose
    listId/id matches the one we already know from the search result."""
    if _depth > 12:
        return None
    if isinstance(obj, dict):
        node_id = obj.get("listId") or obj.get("id")
        if node_id is not None and str(node_id) == str(list_id):
            return obj
        for v in obj.values():
            found = find_ad_node(v, list_id, _depth + 1)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = find_ad_node(item, list_id, _depth + 1)
            if found is not None:
                return found
    return None


def fetch_listing_detail(session, url, list_id, debug_dump=False):
    """Fetch a single listing's own page and return the ad's JSON node (or
    None on any failure) - same __NEXT_DATA__ technique as the search
    pages, just one request per listing. Prints a reason to stderr on
    failure so --fetch-details doesn't fail silently."""
    try:
        r = session.get(url, timeout=20)
    except Exception as e:
        print(f"  [detail] request failed for {url}: {e}", file=sys.stderr)
        return None
    if r.status_code != 200:
        print(f"  [detail] {url}: HTTP {r.status_code}", file=sys.stderr)
        if debug_dump:
            with open("debug_detail_page.html", "w") as f:
                f.write(r.text)
            print("  -> dumped raw HTML to debug_detail_page.html", file=sys.stderr)
        return None
    m = NEXT_DATA_RE.search(r.text)
    if not m:
        print(f"  [detail] {url}: no __NEXT_DATA__ found in response", file=sys.stderr)
        if debug_dump:
            with open("debug_detail_page.html", "w") as f:
                f.write(r.text)
            print("  -> dumped raw HTML to debug_detail_page.html", file=sys.stderr)
        return None
    try:
        nd = json.loads(m.group(1))
    except Exception as e:
        print(f"  [detail] {url}: could not parse __NEXT_DATA__ ({e})", file=sys.stderr)
        if debug_dump:
            with open("debug_detail_page_raw.json", "w") as f:
                f.write(m.group(1))
            print("  -> dumped raw (unparsed) JSON to debug_detail_page_raw.json", file=sys.stderr)
        return None
    if debug_dump:
        with open("debug_detail_page.json", "w") as f:
            json.dump(nd, f, indent=2)
        print("  -> dumped a full detail-page __NEXT_DATA__ to debug_detail_page.json", file=sys.stderr)
    node = find_ad_node(nd, list_id)
    if node is None:
        print(f"  [detail] {url}: __NEXT_DATA__ parsed fine but no node matched listId={list_id}", file=sys.stderr)
    return node


def enrich_with_details(session, rows, limit, debug_dump):
    """For each row missing any of DETAIL_FILL_COLUMNS, visit its listing
    page once and fill in whatever the detail page has that the search
    page didn't. Mutates rows in place."""
    targets = [r for r in rows if any(not r.get(col) for col in DETAIL_FILL_COLUMNS)]
    if limit:
        targets = targets[:limit]
    total = len(targets)
    for i, row in enumerate(targets, 1):
        url = row.get("Listing URL", "")
        id_m = re.search(r"(\d+)(?:\.\w+)?/?$", url)
        list_id = id_m.group(1) if id_m else url.rstrip("/").rsplit("/", 1)[-1]
        node = fetch_listing_detail(session, url, list_id, debug_dump and i == 1)
        if node:
            # The matched node is the outer ad wrapper ({"id", "type",
            # "attributes", ...}), same shape as the search-results items -
            # the actual fields (body, tenure, etc) live one level down in
            # "attributes", not on the wrapper itself.
            attrs = node.get("attributes", node) if isinstance(node, dict) else node
            for col, key in DETAIL_FILL_COLUMNS.items():
                if not row.get(col):
                    val = field_value(attrs, key)
                    if key == "listed_at":
                        val = format_date_ddmmyyyy(val)
                    elif key == "description":
                        val = clean_description(val)
                    if val:
                        row[col] = val
        print(f"[details] {i}/{total}: {url}", file=sys.stderr)
        polite_sleep(1.0, 2.0)


def scrape_category(session, label, base_url, max_pages, owner_only, debug_dump, asset_type):
    results = []
    for pg in range(1, max_pages + 1):
        url = base_url if pg == 1 else f"{base_url}&o={pg}"
        try:
            r = session.get(url, timeout=20)
        except Exception as e:
            print(f"[{label}] page {pg}: request failed: {e}", file=sys.stderr)
            break

        if r.status_code == 429:
            print(f"[{label}] page {pg}: rate limited, waiting 60s", file=sys.stderr)
            time.sleep(60)
            continue
        if r.status_code != 200:
            print(f"[{label}] page {pg}: HTTP {r.status_code}", file=sys.stderr)
            break

        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', r.text, re.DOTALL)
        if not m:
            print(f"[{label}] page {pg}: no __NEXT_DATA__ found", file=sys.stderr)
            if debug_dump:
                with open(f"debug_{label}_p{pg}.html", "w") as f:
                    f.write(r.text)
                print(f"  -> dumped raw HTML to debug_{label}_p{pg}.html", file=sys.stderr)
            break

        try:
            nd = json.loads(m.group(1))
            store = nd["props"]["pageProps"]["initialStore"]
            ads = store.get("ads", []) + store.get("featuredAds", [])
        except Exception as e:
            print(f"[{label}] page {pg}: could not parse __NEXT_DATA__ ({e})", file=sys.stderr)
            if debug_dump:
                with open(f"debug_{label}_p{pg}.json", "w") as f:
                    f.write(m.group(1))
                print(f"  -> dumped raw JSON to debug_{label}_p{pg}.json", file=sys.stderr)
            break

        if not ads:
            print(f"[{label}] page {pg}: 0 ads - stopping", file=sys.stderr)
            break

        if debug_dump and pg == 1 and ads:
            with open(f"debug_{label}_first_ad.json", "w") as f:
                json.dump(ads[0], f, indent=2)
            print(f"  -> dumped first ad's full structure to debug_{label}_first_ad.json", file=sys.stderr)

        page_count = 0
        for item in ads:
            a = item.get("attributes", item)
            inner = a.get("agentData") or {}
            inner = inner.get("data", inner) if isinstance(inner, dict) else {}
            inner = inner if isinstance(inner, dict) else {}

            if owner_only and not is_owner_listing(inner):
                continue

            phone = norm_phone(a.get("phone") or a.get("sellerPhone"))
            agency = inner.get("storeParamsCompanyName") or inner.get("storeName") or None
            has_wa = bool(inner.get("userParamsEnableWa")) if isinstance(inner, dict) else False

            params = extract_category_params(a)
            results.append({
                "Title":         build_title(a),
                "Hero Image URL": extract_hero_image_url(item),
                "Category":      build_category(label, a),
                "Listed Date":   format_date_ddmmyyyy(_first(a, MUDAH_FIELD_CANDIDATES["listed_at"])),
                "Price":         format_price(_first(a, MUDAH_FIELD_CANDIDATES["price"])),
                "Asset Type":    classify_asset_type(field_value(a, "property_type", params), asset_type),
                "Property Type": field_value(a, "property_type", params),
                "Location":      _first(a, MUDAH_FIELD_CANDIDATES["location"]),
                "State":         _first(a, MUDAH_FIELD_CANDIDATES["state"]),
                "Tenure":        field_value(a, "tenure", params),
                "Furnishing":    field_value(a, "furnishing", params),
                "Bedrooms":      field_value(a, "bedroom", params),
                "Bathrooms":     field_value(a, "bathroom", params),
                "Size (sqft)":   field_value(a, "size", params),
                "Land Size":     field_value(a, "land_size", params),
                "Description":   clean_description(_first(a, MUDAH_FIELD_CANDIDATES["description"])),
                "Seller Name":   a.get("name") or None,
                "Agency":        agency,
                "Phone":         phone,
                "Has WhatsApp":  has_wa,
                "Listing URL":   build_listing_url(a),
            })
            page_count += 1

        print(f"[{label}] page {pg}: {page_count} matching listings "
              f"({len(results)} total so far)", file=sys.stderr)
        if page_count == 0 and pg > 3:
            break
        polite_sleep()
    return results


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

BAND_FILL_EVEN = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
BAND_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Distinct background per Asset Type so the eye can scan by property
# category at a glance, on top of the zebra row banding.
ASSET_TYPE_FILL = {
    "residential": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
    "commercial":  PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "land":        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}

DATA_ROW_HEIGHT = 18


def style_listings_sheet(ws, df):
    """Bold coloured header row, frozen header + autofilter, a numbered
    "No." column, uniform-height rows with the Description column
    collapsible (Excel column grouping) instead of blowing out row
    height per listing, zebra row banding, a colour-scale gradient on
    Price (RM), and an Asset Type colour tag - readable at a glance
    instead of plain default openpyxl output."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "B2"  # freeze header row + the "No." column
    ws.auto_filter.ref = ws.dimensions

    col_letter_by_name = {}
    desc_col = None
    for i, name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        col_letter_by_name[name] = col_letter
        if name == "No.":
            ws.column_dimensions[col_letter].width = 6
            continue
        length = max((df[name].apply(lambda v: len(str(v)) if pd.notna(v) else 0).max() if len(df) else 0), len(str(name)))
        ws.column_dimensions[col_letter].width = min(length + 2, 60)
        if name == "Description":
            desc_col = col_letter
            ws.column_dimensions[col_letter].width = 60
            # Collapsible via Excel's column outline/grouping: starts
            # collapsed (hidden) with a "+" toggle above the sheet to
            # expand and read the full text on demand, instead of every
            # row being stretched tall to fit it.
            ws.column_dimensions[col_letter].outlineLevel = 1
            ws.column_dimensions[col_letter].hidden = True
    ws.sheet_properties.outlinePr.summaryRight = True

    asset_type_col = col_letter_by_name.get("Asset Type")
    for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[r].height = DATA_ROW_HEIGHT
        band = BAND_FILL_EVEN if r % 2 == 0 else BAND_FILL_ODD
        asset_type_val = ws[f"{asset_type_col}{r}"].value if asset_type_col else None
        for cell in row:
            cell.border = THIN_BORDER
            cell.fill = ASSET_TYPE_FILL.get(asset_type_val, band)
            if desc_col and cell.column_letter == desc_col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    price_col = col_letter_by_name.get("Price (RM)")
    if price_col and len(df):
        rng = f"{price_col}2:{price_col}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )


def build_summary_df(df):
    """One row per Asset Type x Category, plus a grand-total row, and how
    many listings in each have a phone number captured."""
    if df.empty:
        return pd.DataFrame(columns=["Asset Type", "Category", "Listings", "With Phone"])

    grouped = (
        df.groupby(["Asset Type", "Category"])
        .agg(Listings=("Listing URL", "count"), **{"With Phone": ("Phone", "count")})
        .reset_index()
        .sort_values(["Asset Type", "Category"])
    )
    total_row = pd.DataFrame([{
        "Asset Type": "TOTAL",
        "Category": "",
        "Listings": len(df),
        "With Phone": int(df["Phone"].notna().sum()),
    }])
    return pd.concat([grouped, total_row], ignore_index=True)


# Areas the user is specifically watching. Note: Seri Tanjung Pinang isn't
# its own Location value in mudah.my's data - those listings show up
# tagged under Tanjong Tokong (STP is a development within that subarea),
# so it's already covered here. The title/description-mention check below
# is just a safety net in case that ever changes.
FOCUS_AREAS = ["Tanjong Tokong", "Tanjung Bungah", "Georgetown"]


def build_focus_area_df(df):
    """Subset of df limited to FOCUS_AREAS by Location, plus any listing
    whose title/description explicitly mentions Seri Tanjung Pinang."""
    if df.empty or "Location" not in df.columns:
        return df.iloc[0:0].copy()
    stp_mask = (
        df["Title"].fillna("").str.contains("Seri Tanjung Pinang", case=False)
        | df["Description"].fillna("").str.contains("Seri Tanjung Pinang", case=False)
    )
    mask = df["Location"].isin(FOCUS_AREAS) | stp_mask
    return df[mask].copy()


def build_summary_by_area_df(df):
    """Pivot table: one row per Location, one column per Asset Type, plus
    a Total column and a grand-total row - sorted by total listings
    descending so the busiest areas show up first."""
    if df.empty:
        return pd.DataFrame(columns=["Location", "Total"])
    pivot = pd.pivot_table(
        df, index="Location", columns="Asset Type", values="Listing URL",
        aggfunc="count", fill_value=0,
    )
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    total_row = pivot.sum(axis=0)
    total_row.name = "TOTAL"
    pivot = pd.concat([pivot, total_row.to_frame().T])
    return pivot.reset_index().rename(columns={"index": "Location"})


def style_summary_sheet(ws, df):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for i, name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        length = max((df[name].apply(lambda v: len(str(v)) if pd.notna(v) else 0).max() if len(df) else 0), len(str(name)))
        ws.column_dimensions[col_letter].width = min(length + 4, 40)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)


def extract_list_id(url):
    m = re.search(r"(\d+)(?:\.\w+)?/?$", url or "")
    return m.group(1) if m else None


def load_seen_registry(path):
    """{listId: first-seen-date-string} persisted across runs so a daily
    cron/launchd job can tell which listings are genuinely new instead of
    re-reporting the same still-live ad every day."""
    try:
        with open(path) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_seen_registry(path, registry):
    with open(path, "w") as f:
        json.dump(registry, f, indent=2)


def apply_dedup_tracking(df, seen_file):
    """Add "First Seen" / "Is New Today" columns by checking each
    listing's ID against the persisted seen-file registry, then update
    the registry with any newly-seen IDs. Mutates and returns df."""
    registry = load_seen_registry(seen_file)
    today = datetime.date.today().strftime("%d/%m/%Y")
    first_seen, is_new = [], []
    for url in df["Listing URL"]:
        list_id = extract_list_id(url)
        if list_id and list_id in registry:
            first_seen.append(registry[list_id])
            is_new.append(False)
        else:
            if list_id:
                registry[list_id] = today
            first_seen.append(today)
            is_new.append(True)
    df["First Seen"] = first_seen
    df["Is New Today"] = is_new
    save_seen_registry(seen_file, registry)
    return df


def price_to_number(raw):
    """Numeric twin of the formatted "RM x,xxx,xxx" Price column, so
    Excel's autofilter can do real numeric filtering/sorting/colour-scale
    conditional formatting instead of alphabetical text comparison."""
    if not raw:
        return None
    digits = re.sub(r"[^\d.]", "", str(raw))
    return float(digits) if digits else None


def sync_to_gsheet(df, gsheet_id, gsheet_key, tab_name):
    """Push a DataFrame straight into a tab of a live Google Sheet (create
    the tab if it doesn't exist yet, otherwise wipe and replace its
    contents with this run's data). Requires a service-account JSON key
    that's been shared with Editor access on the target sheet - see the
    --gsheet-key/--gsheet-id help text.

    Uses value_input_option="RAW" rather than "USER_ENTERED" - the latter
    auto-parses numeric-looking strings into actual numbers, which would
    silently strip the leading zero off every phone number (the same
    corruption that hit the Excel output early on in this project, from
    Excel doing the same kind of auto-conversion)."""
    try:
        import gspread
    except ImportError:
        print("Missing dependency for --gsheet-id. Run: pip install gspread google-auth", file=sys.stderr)
        return

    try:
        gc = gspread.service_account(filename=gsheet_key)
        sh = gc.open_by_key(gsheet_id)
    except Exception as e:
        print(f"Could not connect to Google Sheet {gsheet_id}: {e}", file=sys.stderr)
        return

    try:
        ws = sh.worksheet(tab_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=max(len(df) + 10, 100), cols=max(len(df.columns) + 2, 20))

    safe_df = df.where(pd.notna(df), "")
    values = [list(safe_df.columns)] + safe_df.values.tolist()
    try:
        ws.update(values, value_input_option="RAW")
        ws.freeze(rows=1)
        ws.format("1:1", {
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "backgroundColor": {"red": 0.12, "green": 0.31, "blue": 0.47},
        })
        print(f"Synced {len(df)} rows to Google Sheet tab '{tab_name}'", file=sys.stderr)
    except Exception as e:
        print(f"Failed to sync tab '{tab_name}' to Google Sheet: {e}", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--pages", type=int, default=50)
    ap.add_argument("--output", default="penang_owners.xlsx")
    ap.add_argument("--owner-only", action="store_true", default=True)
    ap.add_argument("--include-agents", dest="owner_only", action="store_false",
                     help="Disable the by-owner filter and keep agent listings too.")
    ap.add_argument("--statewide", action="store_true",
                     help="Scrape all of Penang state instead of just Penang Island (e.g. "
                          "the ~1858-listing 'By Owner' search at mudah.my/penang/properties-for-sale).")
    ap.add_argument("--debug-dump", action="store_true",
                     help="Dump raw HTML/JSON for the first page/ad so field names can be verified.")
    ap.add_argument("--fetch-details", action="store_true",
                     help="Also visit each listing's own page to fill in Description and any "
                          "other detail fields missing from the search-results JSON. Slower "
                          "(one extra request per listing).")
    ap.add_argument("--detail-limit", type=int, default=0,
                     help="With --fetch-details, cap how many listings get the detail-page visit "
                          "(0 = all).")
    ap.add_argument("--seen-file", default="penang_seen_listings.json",
                     help="Path to a JSON registry of previously-seen listing IDs, used to flag "
                          "which listings are new since the last run (for daily scheduled runs, "
                          "point this at the same path every time so state persists).")
    ap.add_argument("--no-dedup-tracking", dest="dedup_tracking", action="store_false", default=True,
                     help="Disable new-vs-already-seen tracking against --seen-file.")
    ap.add_argument("--gsheet-id", default=None,
                     help="Google Sheet ID to push results to live (the long ID in the sheet's "
                          "URL between /d/ and /edit). Requires --gsheet-key. Needs: "
                          "pip install gspread google-auth")
    ap.add_argument("--gsheet-key", default=None,
                     help="Path to a Google service-account JSON credentials file with Editor "
                          "access on --gsheet-id (share the sheet with the service account's "
                          "client_email once, from the JSON key file).")
    args = ap.parse_args()

    try:
        from curl_cffi import requests as cffi_requests
    except ImportError:
        print("Missing dependency. Run: pip install curl_cffi", file=sys.stderr)
        sys.exit(1)

    session = cffi_requests.Session(impersonate="chrome124")
    session.headers.update({
        "Accept-Language": "en-MY,en;q=0.9,ms;q=0.8",
        "Referer": "https://www.mudah.my/",
    })

    all_results = []
    for label, slug, asset_type in CATEGORIES:
        if args.statewide:
            base_url = f"https://www.mudah.my/penang/{slug}?adsby=false"
        else:
            base_url = f"https://www.mudah.my/malaysia/{slug}?adsby=false&subarea={PENANG_SUBAREAS}"
        rows = scrape_category(session, label, base_url, args.pages, args.owner_only,
                                args.debug_dump, asset_type)
        all_results.extend(rows)

    if not all_results:
        print(
            "No listings extracted. Try --debug-dump to capture raw output, or the "
            "__NEXT_DATA__ shape may differ from what this script expects.",
            file=sys.stderr,
        )
        sys.exit(1)

    if args.fetch_details:
        print(f"Fetching listing detail pages to fill in Description/etc "
              f"({'all' if not args.detail_limit else args.detail_limit} of {len(all_results)} listings)...",
              file=sys.stderr)
        enrich_with_details(session, all_results, args.detail_limit, args.debug_dump)

    df = pd.DataFrame(all_results)
    df.drop_duplicates(subset=["Listing URL"], inplace=True)

    df.insert(df.columns.get_loc("Price") + 1, "Price (RM)", df["Price"].apply(price_to_number))

    new_count = None
    if args.dedup_tracking:
        df = apply_dedup_tracking(df, args.seen_file)
        new_count = int(df["Is New Today"].sum())
        print(f"{new_count}/{len(df)} listings are new since the last run "
              f"(seen-file: {args.seen_file})", file=sys.stderr)

    df.insert(0, "No.", range(1, len(df) + 1))
    summary_df = build_summary_df(df)
    summary_by_area_df = build_summary_by_area_df(df)
    focus_df = build_focus_area_df(df)
    if len(focus_df):
        focus_df["No."] = range(1, len(focus_df) + 1)

    new_df = None
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All Listings")
        style_listings_sheet(writer.sheets["All Listings"], df)

        if args.dedup_tracking:
            new_df = df[df["Is New Today"]].copy()
            new_df["No."] = range(1, len(new_df) + 1)
            new_df.to_excel(writer, index=False, sheet_name="New Listings")
            style_listings_sheet(writer.sheets["New Listings"], new_df)

        focus_df.to_excel(writer, index=False, sheet_name="Focus Area")
        style_listings_sheet(writer.sheets["Focus Area"], focus_df)

        summary_by_area_df.to_excel(writer, index=False, sheet_name="Summary by Area")
        style_summary_sheet(writer.sheets["Summary by Area"], summary_by_area_df)

        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        style_summary_sheet(writer.sheets["Summary"], summary_df)

    print(f"Saved {len(df)} unique listings to {args.output}")
    with_phone = df["Phone"].notna().sum()
    print(f"  {with_phone}/{len(df)} have a phone number extracted directly from the page JSON")
    print(f"  {len(focus_df)} listings in the Focus Area tab (Tanjong Tokong / Tanjung Bungah / Georgetown)")
    if new_count is not None:
        print(f"  {new_count} listings are new since the last run - see the 'New Listings' sheet")

    if args.gsheet_id and args.gsheet_key:
        sync_to_gsheet(df, args.gsheet_id, args.gsheet_key, "All Listings")
        if new_df is not None:
            sync_to_gsheet(new_df, args.gsheet_id, args.gsheet_key, "New Listings")
        sync_to_gsheet(focus_df, args.gsheet_id, args.gsheet_key, "Focus Area")
        sync_to_gsheet(summary_by_area_df, args.gsheet_id, args.gsheet_key, "Summary by Area")
        sync_to_gsheet(summary_df, args.gsheet_id, args.gsheet_key, "Summary")
    elif args.gsheet_id or args.gsheet_key:
        print("Both --gsheet-id and --gsheet-key are required to sync to Google Sheets - skipping sync.",
              file=sys.stderr)


if __name__ == "__main__":
    main()
